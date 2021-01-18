# -*- coding: utf-8 -*-
"""
Created on Mon Jan 11 09:26:23 2021

@author: Kolovos
"""

#--FORMATING using 'openpyxl' 
# Writing from python to .xlsx
#copy from 'OutputExampleOP.xlsx' to 'ProjectResults.xlsx'---
import openpyxl

# ----------------'WRITING' on Excel 'EcoAnResuls.xlsx'----------------------
# create blank Workbook object & active sheet 
wbd = openpyxl.Workbook()     # wb destination
page=wbd.active               # page @ destination
# page titles:
page.title = 'EcoAnResuls'

#-------load 'OutputExampleOP.xlsx'---------------------------------------------
wbs2= openpyxl.load_workbook('OutputExampleOP.xlsx') # wb source2
sheet2 = wbs2.active 

#---from 'OutputExampleOP.xlsx' to 'EcoAnResuls.xlsx'--------------------------
# total number of rows & columns in 'OutputExampleOP.xlsx'----------------
Rm = sheet2.max_row 
Cm = sheet2.max_column 
  
# copying the cell values from source 'OutputExampleOP.xlsx' 
for i in range (1, Rm + 1): 
    for j in range (1, Cm + 1):  
        Cell =sheet2.cell(row = i, column = j).value 
  
        # Transfer value to destination 'EcoAnResuls.xlsx' 
        page.cell(row = i, column = j).value = Cell 
#-------------------------------FORMATTING------------------------------------
#--import 'styling' tools-----------------------------------------------------
from openpyxl.styles import Font                # style page
from openpyxl.styles.borders import Border,Side # Border
import copy                                     # Wrap, Center..

# Bold --------------------------------------------------
Bold = Font(bold=True)
page['A1'].font = Bold
page['D1'].font = Bold
page['A2'].font = Bold
page['A25'].font = Bold
page['D25'].font = Bold
page['A26'].font = Bold
# Merge--------------------------------------------------
page.merge_cells('E2:I2')
page.merge_cells('E26:I26')
# 2 decimal places for all colmns indicated-----------
for row in range(1,Rm+1):
    page["E{}".format(row)].number_format = '#,##0.00'
    page["F{}".format(row)].number_format = '#,##0.00'
    page["G{}".format(row)].number_format = '#,##0.00'
    page["H{}".format(row)].number_format = '#,##0.00'
    page["I{}".format(row)].number_format = '#,##0.00'
# Wrap_text & Center-----------------------------------------
for row in page.iter_rows():
    for cell in row:      
        alignment = copy.copy(cell.alignment)
        alignment.wrapText=True
        alignment.horizontal = 'center'
        alignment.vertical = 'center'
        cell.alignment = alignment
# Border------------------------------------------------------
LeftThck = Border(left=Side(style='thick'))    # col E
RightThck = Border(right=Side(style='thick'))  # col I
TopThck = Border(top=Side(style='thick'))      # row 14,38
BottThck = Border(bottom=Side(style='thick'))  # row 3,27
# row for existing Wind                     
for row in page.iter_rows(min_row=4, max_row=13, min_col=5, max_col=5):
    for cell in row:
        cell.border=LeftThck
for row in page.iter_rows(min_row=4, max_row=13, min_col=9, max_col=9):
    for cell in row:
        cell.border=RightThck
for row in page.iter_rows(min_row=3, max_row=3, min_col=5, max_col=9):
    for cell in row:
        cell.border=BottThck
for row in page.iter_rows(min_row=14, max_row=14, min_col=5, max_col=9):
    for cell in row:
        cell.border=TopThck
# row for existing Solar                     
for row in page.iter_rows(min_row=28, max_row=37, min_col=5, max_col=5):
    for cell in row:
        cell.border=LeftThck
for row in page.iter_rows(min_row=28, max_row=37, min_col=9, max_col=9):
    for cell in row:
        cell.border=RightThck
for row in page.iter_rows(min_row=27, max_row=27, min_col=5, max_col=9):
    for cell in row:
        cell.border=BottThck
for row in page.iter_rows(min_row=38, max_row=38, min_col=5, max_col=9):
    for cell in row:
        cell.border=TopThck

wbd.save('EcoAnResulsM3.xlsx')