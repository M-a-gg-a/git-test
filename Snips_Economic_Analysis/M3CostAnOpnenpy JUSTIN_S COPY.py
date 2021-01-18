# -*- coding: utf-8 -*-
"""
Created on Sun Jan  3 05:29:54 2021

@author: Kolovos
"""

#-------------Input from M1------------------------------------
'''
# --------------------------Assumptions------------------------------
1. GIVES A YEAR FORCAST (main reason missing gas frwd curve in farther future)
 . Also the 'Market values' & 'weather patterns' in M2 will have to be for multiple years 
2. values  given in 'Input from M2' section MUST be initialized in prior modules
 . preferably M1 (all var = 0)
3. M1 should start with 'ExistingWind = 1' if existing resource is Wind (0 ir Solar)
'''
# At M1 we choose if we start with Existing Wind (=1) of not (Solar = 0)
ExistingWind = 0


'''
------'Key' for numbering/labeling additional resources and combinations-------

Additional resources
1 = Wind 
2 = Solar
3 = Battery Wind
4 = Battery Solar
5 = Combine Cycle (CC)
6 = Combination Tutbine (CT)
7 = RIC
8 = Aeroderivative (A)
2 conmbinations= double #
3 conmbinations= triple #    (not used in senarios here)
4 conmbinations= quadruple # (not used in senarios here)
'''
#-----------------------------Input from M2------------------------------------
# Create Test Variables (outcomes of M2)
# these results should arive from M2 (here are just for testing)
    
#MWh in Year for Solar 
MWhS = 500
# curtailment additions [MWh] Solar in a Year
MWhS_crt = 45   # used by battery when available - Solar
MWhS_Xcrt = 5   # Extra curtailment not stored - Solar
MWhS_Tcrt = 50  # Total (used & stored) - Solar
 
#MWh in Year for Wind
MWhW = 550
# curtailment additions [MWh] Wind in a Year
MWhW_crt = 50   # used by battery when available - Wind
MWhW_Xcrt = 10  # Extra curtailment not stored - Wind
MWhW_Tcrt = 60  # Total (used & stored) - Wind
    
# Values from M2 independed of Existing Resource  
# CAPITAL CC, CT, RIC, A are used when ONLY additional resource   
# CC MWh produced each month
JrCC = 30; FbCC= 15; MrCC = 40; ApCC =70; MyCC = 35;JnCC = 20;
JlCC = 70; AgCC=60; SpCC = 35;OcCC = 60; NvCC = 10; DcCC = 50
# CT MWh produced each month
JrCT = 0; FbCT= 0; MrCT = 0; ApCT =0; MyCT = 0;JnCT = 0;
JlCT = 0; AgCT=0; SpCT = 0;OcCT = 0; NvCT = 0; DcCT = 0
# RIC MWh produced each month
JrRIC = 30; FbRIC= 15; MrRIC = 0; ApRIC =70; MyRIC = 35;JnRIC = 20;
JlRIC = 70; AgRIC=60; SpRIC = 35;OcRIC = 60; NvRIC = 0; DcRIC = 50
# A MWh produced each month
JrA = 0; FbA= 0; MrA = 0; ApA = 0; MyA = 0;JnA = 0;
JlA = 0; AgA= 0; SpA = 0;OcA = 0; NvA = 00; DcA = 0

# small Cc, Ct, Ric _a are used when COMBINED additional resource   
# Cc MWh produced each month
JrCc = 15; FbCc= 10; MrCc = 20; ApCc =35; MyCc = 20;JnCc = 10;
JlCc = 35; AgCc=30; SpCc = 20;OcCc = 10; NvCc = 5; DcCc = 25
# Ct MWh produced each month
JrCt = 0; FbCt= 0; MrCt = 0; ApCt =0; MyCt = 0;JnCt = 0;
JlCt = 0; AgCt=0; SpCt = 0;OcCt = 0; NvCt = 0; DcCt = 0
# Ric MWh produced each month
JrRic = 15; FbRic= 10; MrRic = 0; ApRic =35; MyRic = 20;JnRic = 10;
JlRic = 35; AgRic=15; SpRic = 20;OcRic = 30; NvRic = 0; DcRic = 25
# _a MWh produced each month
Jr_a = 0; Fb_a= 0; Mr_a = 0; Ap_a = 0; My_a = 0;Jn_a = 0;
Jl_a = 0; Ag_a= 0; Sp_a = 0;Oc_a = 0; Nv_a = 0; Dc_a = 0
 
#***************************Starting M3***************************************  
#--FORMATING using 'openpyxl' copy from 'OutputExampleOP.xlsx' to 'ProjectResults.xlsx'---
# Excel-Python using 'openpyxl'
import openpyxl

# ----------------'WRITING' on Excel 'EcoAnResuls.xlsx'----------------------
# create blank Workbook object & active sheet 
wbd = openpyxl.Workbook()     # wb destination
page=wbd.active               # page @ destination
# page titles:
page.title = 'EcoAnResuls'

#-------load 'OutputExampleOP.xlsx'---------------------------------------------
wbs2= openpyxl.load_workbook('OutputExampleOP.xlsx')
sheet2 = wbs2.active 

#---from 'OutputExampleOP.xlsx' to 'EcoAnResuls.xlsx'--------------------------
# total number of rows & columns in 'OutputExampleOP.xlsx'----------------
Rm = sheet2.max_row 
Cm = sheet2.max_column 
  
# copying the cell values from source 'OutputExampleOP.xlsx' 
for i in range (1, Rm + 1): 
    for j in range (1, Cm + 1):  
        Cell =sheet2.cell(row = i, column = j).value 
  
        # Transfer value to destination 'EcoAnResuls.xlsx' 
        page.cell(row = i, column = j).value = Cell 
#-------------------------------FORMATTING------------------------------------
#--import 'styling' tools-----------------------------------------------------
from openpyxl.styles import Font                # style page
from openpyxl.styles.borders import Border,Side # Border
import copy                                     # Wrap, Center..

# Bold --------------------------------------------------
Bold = Font(bold=True)
page['A1'].font = Bold
page['D1'].font = Bold
page['A2'].font = Bold
page['A25'].font = Bold
page['D25'].font = Bold
page['A26'].font = Bold
# Merge--------------------------------------------------
page.merge_cells('E2:I2')
page.merge_cells('E26:I26')
# 2 decimal places----------------------------------------
for row in range(1,Rm+1):
    page["E{}".format(row)].number_format = '#,##0.00'
    page["F{}".format(row)].number_format = '#,##0.00'
    page["G{}".format(row)].number_format = '#,##0.00'
    page["H{}".format(row)].number_format = '#,##0.00'
    page["I{}".format(row)].number_format = '#,##0.00'
# Wrap_text & Center-----------------------------------------
for row in page.iter_rows():
    for cell in row:      
        alignment = copy.copy(cell.alignment)
        alignment.wrapText=True
        alignment.horizontal = 'center'
        alignment.vertical = 'center'
        cell.alignment = alignment
# Border------------------------------------------------------
LeftThck = Border(left=Side(style='thick'))    # col E
RightThck = Border(right=Side(style='thick'))  # col I
TopThck = Border(top=Side(style='thick'))      # row 14,38
BottThck = Border(bottom=Side(style='thick'))  # row 3,27
# row for existing Wind                     
for row in page.iter_rows(min_row=4, max_row=13, min_col=5, max_col=5):
    for cell in row:
        cell.border=LeftThck
for row in page.iter_rows(min_row=4, max_row=13, min_col=9, max_col=9):
    for cell in row:
        cell.border=RightThck
for row in page.iter_rows(min_row=3, max_row=3, min_col=5, max_col=9):
    for cell in row:
        cell.border=BottThck
for row in page.iter_rows(min_row=14, max_row=14, min_col=5, max_col=9):
    for cell in row:
        cell.border=TopThck
# row for existing Solar                     
for row in page.iter_rows(min_row=28, max_row=37, min_col=5, max_col=5):
    for cell in row:
        cell.border=LeftThck
for row in page.iter_rows(min_row=28, max_row=37, min_col=9, max_col=9):
    for cell in row:
        cell.border=RightThck
for row in page.iter_rows(min_row=27, max_row=27, min_col=5, max_col=9):
    for cell in row:
        cell.border=BottThck
for row in page.iter_rows(min_row=38, max_row=38, min_col=5, max_col=9):
    for cell in row:
        cell.border=TopThck
#*************************Actual Code Starts here***************************
#------------------List yearMWh =[monthMWh,monthMWh... ]--------------------
# ---------------------When gas ONLY additional resource--------------------
# CC MWh in Year
MWhCC =[JrCC,FbCC,MrCC,ApCC,MyCC,JnCC,         # this is the val. of the list
        JlCC,AgCC,SpCC,OcCC,NvCC,DcCC]
# CT MWh in Year
MWhCT =[JrCT,FbCT,MrCT,ApCT,MyCT,JnCT,         # this is the val. of the list
        JlCT,AgCT,SpCT,OcCT,NvCT,DcCT]
# RIC MWh in Year
MWhRIC =[JrRIC,FbRIC,MrRIC,ApRIC,MyRIC,JnRIC,  # this is the val. of the list
        JlRIC,AgRIC,SpRIC,OcRIC,NvRIC,DcRIC]
# A MWh in Year
MWhA =[JrA,FbA,MrA,ApA,MyA,JnA,                # this is the val. of the list
        JlA,AgA,SpA,OcA,NvA,DcA]

# ---------------------When gas COMBINED additional resource-------------------
# Cc MWh in Year
MWhCc =[JrCc,FbCc,MrCc,ApCc,MyCc,JnCc,         # this is the val. of the list
        JlCc,AgCc,SpCc,OcCc,NvCc,DcCc]
# Ct MWh in Year
MWhCt =[JrCt,FbCt,MrCt,ApCt,MyCt,JnCt,         # this is the val. of the list
        JlCt,AgCt,SpCt,OcCt,NvCt,DcCt]
# Ric MWh in Year
MWhRic =[JrRic,FbRic,MrRic,ApRic,MyRic,JnRic,  # this is the val. of the list
        JlRic,AgRic,SpRic,OcRic,NvRic,DcRic]
# _a MWh in Year
MWh_a =[Jr_a,Fb_a,Mr_a,Ap_a,My_a,Jn_a,         # this is the val. of the list
        Jl_a,Ag_a,Sp_a,Oc_a,Nv_a,Dc_a]

#--------------------------Formulas------------------------------------------
# ----------'Reading' from ecxel 'EconomicAnalysisM3.xlsx'-------------------
# pandas can read formulas directly from excel sheet
import pandas as pd

sheet = pd.read_excel('EconomicAnalysisM3.xlsx') 

# assign values to var in [$/MWh] from 'EconomicAnalysisM3.xlsx'
C26 = sheet.iloc[24, 2]
D27 = sheet.iloc[25, 3]
E26 = sheet.iloc[24, 4]
E27= sheet.iloc[25, 4]

#RESOURCE---RESOURCE---RESOURCE---RESOURCE---RESOURCE---RESOURCE------------- 
#----------------------Share Senarios (All Existing Resources)---------------
#-----------Calculating gas values when gas ONLY additional resource------
# 5 Combine Cycle
# initialization for the 'for loops'
CC_DpMWh = 0   # from source excel (fromula) = $/MWh
CCmonth_MWh= 0 # MWh for each month -LIST
CC_monthD= 0   # $ for each month 
CC_D=0         # Total $ CC
SumMWhCC= 0    # Total MWh CC
# formula =(B29*$F$9)/1000+$F$7
F9= sheet.iloc[7,5]  # from 'EconomicAnalysisM3.xlsx'
F7= sheet.iloc[5,5]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] CC
for i in range (27,39):  
    CC_DpMWh= (sheet.iloc[i,1]* F9/1000) + F7
    CCmonth_MWh= MWhCC[i-27]  
    CC_monthD = CC_DpMWh * CCmonth_MWh 
    SumMWhCC+= CCmonth_MWh    # Total [MWH]
    CC_D+=CC_monthD           # Total [$]
# Total $/MWh for all months
if SumMWhCC > 0:
    CC_DpE=CC_D/SumMWhCC
else:
     CC_DpE=0  

# 6 Combination Turbine 
# initialization for the 'for loops'
CT_DpMWh = 0   # from source excel (fromula) = $/MWh
CTmonth_MWh= 0 # MWh for each month
CT_monthD= 0   # $ for each month ($)
CT_D=0         # Total $ CT
SumMWhCT= 0    # Total MWh CT
# formula =(B29*$G$9)/1000+$G$7
G9= sheet.iloc[7,6]  # from 'EconomicAnalysisM3.xlsx'
G7= sheet.iloc[5,6]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] CT
for i in range (27,39):  
    CT_DpMWh= (sheet.iloc[i,1]* G9/1000) + G7
    CTmonth_MWh= MWhCT[i-27]  
    CT_monthD = CT_DpMWh * CTmonth_MWh
    SumMWhCT+= CTmonth_MWh    # Total [MWH]
    CT_D+=CT_monthD           # Total [$]
if SumMWhCT > 0:
    CT_DpE=CT_D/SumMWhCT
else:
     CT_DpE=0  
      
# 7 Reciprocating IC
# initialization for the 'for loops'
RIC_DpMWh = 0   # from source excel (fromula) = $/MWh
RICmonth_MWh= 0 # MWh for each month
RIC_monthD= 0   # $ for each month ($)
RIC_D=0           # Total $ RIC
SumMWhRIC= 0    # Total MWh RIC
# formula =(B29*$H$9)/1000+$H$7
H9= sheet.iloc[7,7]  # from 'EconomicAnalysisM3.xlsx'
H7= sheet.iloc[5,7]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] RIC
for i in range (27,39):  
    RIC_DpMWh= (sheet.iloc[i,1]* H9/1000) + H7
    RICmonth_MWh= MWhRIC[i-27]  
    RIC_monthD = RIC_DpMWh * RICmonth_MWh
    SumMWhRIC+= RICmonth_MWh    # Total [MWH]
    RIC_D+=RIC_monthD           # Total [$]
if SumMWhRIC > 0:
    RIC_DpE=RIC_D/SumMWhRIC
else:
     RIC_DpE=0  
 
# 8 Aeroderivative
# initialization for the 'for loops'
A_DpMWh = 0   # from source excel (fromula) = $/MWh
Amonth_MWh= 0 # MWh for each month
A_monthD= 0   # $ for each month ($)
A_D=0         # Total $ A
SumMWhA= 0    # Total MWh A
# formula =(B29*$I$9)/1000+$I$7 
I9= sheet.iloc[7,8]  # from 'EconomicAnalysisM3.xlsx'
I7= sheet.iloc[5,8]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] A
for i in range (27,39):  
    A_DpMWh= (sheet.iloc[i,1]* I9/1000) + I7
    Amonth_MWh= MWhA[i-27]  
    A_monthD = A_DpMWh * Amonth_MWh
    SumMWhA+= Amonth_MWh    # Total [MWH]
    A_D+= A_monthD          # Total [$]
if SumMWhA > 0:
    A_DpE=A_D/SumMWhA
else:
     A_DpE=0 

#-----------Calculating gas values when gas COMBINED additional resource------
# 5b. Combine Cycle
# initialization for the 'for loops'
Cc_DpMWh = 0   # from source excel (fromula) = $/MWh
Ccmonth_MWh= 0 # MWh for each month -LIST
Cc_monthD= 0   # $ for each month 
Cc_D=0         # Total $ Cc
SumMWhCc= 0    # Total MWh Cc
# formula =(B29*$F$9)/1000+$F$7
F9= sheet.iloc[7,5]  # from 'EconomicAnalysisM3.xlsx'
F7= sheet.iloc[5,5]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] Cc
for i in range (27,39):  
    Cc_DpMWh= (sheet.iloc[i,1]* F9/1000) + F7
    Ccmonth_MWh= MWhCc[i-27]  
    Cc_monthD = Cc_DpMWh * Ccmonth_MWh 
    SumMWhCc+= Ccmonth_MWh    # Total [MWH]
    Cc_D+=Cc_monthD           # Total [$]
# Total $/MWh for all months
if SumMWhCc > 0:
    Cc_DpE=Cc_D/SumMWhCc
else:
     Cc_DpE=0  

# 6b. Combination Turbine 
# initialization for the 'for loops'
Ct_DpMWh = 0   # from source excel (fromula) = $/MWh
Ctmonth_MWh= 0 # MWh for each month
Ct_monthD= 0   # $ for each month ($)
Ct_D=0         # Total $ Ct
SumMWhCt= 0    # Total MWh Ct
# formula =(B29*$G$9)/1000+$G$7
G9= sheet.iloc[7,6]  # from 'EconomicAnalysisM3.xlsx'
G7= sheet.iloc[5,6]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] Ct
for i in range (27,39):  
    Ct_DpMWh= (sheet.iloc[i,1]* G9/1000) + G7
    Ctmonth_MWh= MWhCt[i-27]  
    Ct_monthD = Ct_DpMWh * Ctmonth_MWh
    SumMWhCt+= Ctmonth_MWh    # Total [MWH]
    Ct_D+=Ct_monthD           # Total [$]
if SumMWhCt > 0:
    Ct_DpE=Ct_D/SumMWhCt
else:
     Ct_DpE=0  
      
# 7b. Reciprocating ic
# initialization for the 'for loops'
Ric_DpMWh = 0   # from source excel (fromula) = $/MWh
Ricmonth_MWh= 0 # MWh for each month
Ric_monthD= 0   # $ for each month ($)
Ric_D=0           # Total $ Ric
SumMWhRic= 0    # Total MWh Ric
# formula =(B29*$H$9)/1000+$H$7
H9= sheet.iloc[7,7]  # from 'EconomicAnalysisM3.xlsx'
H7= sheet.iloc[5,7]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] Ric
for i in range (27,39):  
    Ric_DpMWh= (sheet.iloc[i,1]* H9/1000) + H7
    Ricmonth_MWh= MWhRic[i-27]  
    Ric_monthD = Ric_DpMWh * Ricmonth_MWh
    SumMWhRic+= Ricmonth_MWh    # Total [MWH]
    Ric_D+=Ric_monthD           # Total [$]
if SumMWhRic > 0:
    Ric_DpE=Ric_D/SumMWhRic
else:
     Ric_DpE=0  
 
# 8b. Aeroderivative (_a)
# initialization for the 'for loops'
_a_DpMWh = 0   # from source excel (fromula) = $/MWh
_amonth_MWh= 0 # MWh for each month
_a_monthD= 0   # $ for each month ($)
_a_D=0         # Total $ _a
SumMWh_a= 0    # Total MWh _a
# formula =(B29*$I$9)/1000+$I$7 
I9= sheet.iloc[7,8]  # from 'EconomicAnalysisM3.xlsx'
I7= sheet.iloc[5,8]  # from 'EconomicAnalysisM3.xlsx'
# Total [MWH] & Total [$] _a
for i in range (27,39):  
    _a_DpMWh= (sheet.iloc[i,1]* I9/1000) + I7
    _amonth_MWh= MWh_a[i-27]  
    _a_monthD = _a_DpMWh * _amonth_MWh
    SumMWh_a += _amonth_MWh   # Total [MWH]
    _a_D+= _a_monthD          # Total [$]
if SumMWh_a > 0:
    _a_DpE= _a_D/SumMWh_a
else:
     _a_DpE=0 

#-----------------Senarior differer [Based on 'Existing Resources']---------
#-------Based upon Existing Resources-----MWhW or MWhS = 0 -distinguishable ---

# 1a) Wind (included  Wind_curtail) - NO BATTERY (in use)
Wind_D = C26*MWhW                          # $ Wind
WindCurt_D = C26*MWhW_Tcrt                 # $ Curtailment
if MWhW > 0:                               # Mechanism to prevent /0
    W_DpE = (Wind_D + WindCurt_D)/ MWhW    # $/MWh (NO battery)
else:
     W_DpE =0
    
# 1b) wind (included  Wind_curtail) + BATTERY (in use)
Wind_D = C26*MWhW                          # $ Wind
windCurt_D = C26*MWhW_Xcrt                 # $ Curtailment (reduced)
if MWhW > 0:                               # Mechanism to prevent /0
    w_DpE = (Wind_D + windCurt_D)/ MWhW    # $/MWh (USE battery)
else:
  w_DpE =0  
  
 # 3 Batteries Wind (included  *curtail taken at wind -1b) Senario)
BattW_D = E26*MWhW_crt                    # $ Batteries
if MWhW_crt >0:                           # Mechanism to prevent /0
    BatW_DpE = BattW_D/ MWhW_crt          # $/MWh BatteryWind
else:
  BatW_DpE =0

# 2a) Solar (included  Solar_curtail) - NO BATTERY (in use)
Solar_D = D27*MWhS                          # $ Solar
SolarCurt_D = D27*MWhS_Tcrt                 # $  Curtailment
if MWhS > 0:                                # Mechanism to prevent /0
    Sol_DpE = (Solar_D + SolarCurt_D)/ MWhS # $/MWh (NO battery)
else:
    Sol_DpE = 0 

# 2b) solar (included  Solar_curtail) + BATTERY (in use)
Solar_D = D27*MWhS                           # $ Solar
solarCurt_D = D27*MWhS_Xcrt                  # $ Curtailment (reduced)
if MWhS > 0:                                 # Mechanism to prevent /0
    sol_DpE = (Solar_D + solarCurt_D)/ MWhS  # $/MWh (USE battery)
else:
    sol_DpE = 0 

 # 4 Batteries Solar  (included  *curtail taken at Solar -2b) Senario))
BattS_D = E27*MWhS_crt                    # $ Batteries
if MWhS_crt >0:                           # Mechanism to prevent /0
    BatS_DpE = BattS_D/ MWhS_crt          # $/MWh BatterySolar
else:
    BatS_DpE =0
#-------------------------------------Combinations-------------------------
#---------------------------------From Existing Solar--------------------
# 13 Battery with Wind
if (MWhW and MWhW_crt):        # Creating Senarion (both should be T)
    WBE= MWhW+MWhW_crt         # Energy Resource
    WBD=Wind_D+BattW_D         # $ Resource
    WBCrE=MWhW_Xcrt            # Energy Curt
    WBCrD=windCurt_D           # $ Curt
    WBDpE =(WBD+ WBCrD) /WBE   # $/MWh Resource
else:
    WBE= 0
    WBD=0
    WBCrE=0
    WBCrD=0
    WBDpE=0
    
# 15 Wind & Cc 
if (MWhW and SumMWhCc):       # Creating Senarion (both should be T)         
    WCcE =MWhW+SumMWhCc       # Energy Resource
    WCcD=Wind_D+Cc_D          # $ Resource
    WCcCrE=MWhW_Tcrt          # Energy Curt
    WCcCrD = WindCurt_D       # $ Curt
    WCcDpE=(WCcD+WCcCrD)/WCcE # $/MWh Resource
else:
    WCcE =0
    WCcD=0
    WCcCrE=0
    WCcCrD = 0
    WCcDpE=0
    
# 16 Wind & Ct 
if (MWhW and SumMWhCt):       # Creating Senarion (both should be T)         
    WCtE=MWhW+SumMWhCt        # Energy Resource
    WCtD=Wind_D+Ct_D          # $ Resource
    WCtCrE=MWhW_Tcrt          # Energy Curt
    WCtCrD=WindCurt_D         # $ Curt
    WCtDpE=(WCtD+WCtCrD)/WCtE # $/MWh Resource
else:
    WCtE=0
    WCtD=0
    WCtCrE=0
    WCtCrD=0
    WCtDpE=0
      
# 17 Wind & Ric 
if (MWhW and SumMWhRic):          # Creating Senarion (both should be T)
    WRicE=MWhW+SumMWhRic          # Energy Resource
    WRicD=Wind_D+Ric_D            # $ Resource
    WRicCrE=MWhW_Tcrt             # Energy Curt
    WRicCrD=WindCurt_D            # $ Curt
    WRicDpE=(WRicD+WRicCrD)/WRicE # $/MWh Resource
else:
    WRicE=0
    WRicD=0
    WRicCrE=0
    WRicCrD=0
    WRicDpE=0
      
# 18 Wind & Aro (_a)
if (MWhW and SumMWh_a):           # Creating Senarion (both should be T)
    W_aE=MWhW+SumMWh_a            # Energy Resource
    W_aD=Wind_D+ _a_D             # $ Resource
    W_aCrE= MWhW_Tcrt             # Energy Curt
    W_aCrD= WindCurt_D            # $ Curt
    W_aDpE= (W_aD+W_aCrD)/W_aE    # $/MWh Resource 
else:
    W_aE=0
    W_aD=0
    W_aCrE=0
    W_aCrD=0
    W_aDpE=0
#--------------------------------------------------------------------------
# ------------------------------From Existing Wind-------------------------
# 24 Battery with Solar
if (MWhS and MWhS_crt):           # Creating Senarion (both should be T)
    SBE= MWhS+MWhS_crt            # Energy Resource
    SBD=Solar_D+BattS_D           # $ Resource
    SBCrE=MWhS_Xcrt               # Energy Curt
    SBCrD=solarCurt_D             # $ Curt
    SBDpE =(SBD+ SBCrD) /SBE      # $/MWh Resource     
else:
    SBE= 0
    SBD=0
    SBCrE=0
    SBCrD=0
    SBDpE =0

# 25 Solar & Cc  
if (MWhS and SumMWhCc):           # Creating Senarion (both should be T)   
    SCcE =MWhS+SumMWhCc           # Energy Resource
    SCcD=Solar_D+ Cc_D            # $ Resource
    SCcCrE=MWhS_Tcrt              # Energy Curt         
    SCcCrD = SolarCurt_D          # $ Curt
    SCcDpE=(SCcD+SCcCrD)/SCcE     # $/MWh Resource 
else:
    SCcE =0
    SCcD=0
    SCcCrE=0
    SCcCrD = 0
    SCcDpE=0

# 26 Solar & Ct 
if (MWhS and SumMWhCt):          # Creating Senarion (both should be T)             
    SCtE=MWhS+SumMWhCt           # Energy Resource
    SCtD=Solar_D+Ct_D            # $ Resource
    SCtCrE=MWhS_Tcrt             # Energy Curt 
    SCtCrD=SolarCurt_D           # $ Curt
    SCtDpE=(SCtD+SCtCrD)/SCtE    # $/MWh Resource
else:
    SCtE=0
    SCtD=0
    SCtCrE=0
    SCtCrD=0
    SCtDpE=0

# 27 Solar & Ric
if (MWhS and SumMWhRic):           # Creating Senarion (both should be T)   
    SRicE=MWhS+SumMWhRic           # Energy Resource
    SRicD=Solar_D+Ric_D            # $ Resource
    SRicCrE=MWhS_Tcrt              # Energy Curt 
    SRicCrD=SolarCurt_D            # $ Curt
    SRicDpE=(SRicD+SRicCrD)/SRicE  # $/MWh Resource
else:
    SRicE=0
    SRicD=0
    SRicCrE=0
    SRicCrD=0
    SRicDpE=0

# 28 Solar & Aro (_a)
if (MWhS and SumMWh_a):             # Creating Senarion (both should be T)    
    S_aE=MWhS+SumMWh_a              # Energy Resource
    S_aD=Solar_D+ _a_D              # $ Resource
    S_aCrE= MWhS_Tcrt               # Energy Curt
    S_aCrD= SolarCurt_D             # $ Curt
    S_aDpE= (S_aD+S_aCrD)/S_aE      # $/MWh Resource
else:
    S_aE=0
    S_aD=0
    S_aCrE=0
    S_aCrD=0
    S_aDpE=0
 
# -------write on Excel Sheet 'EcoAnResulsM3.xlsx'-------------------
#-------Resources & combinations from 'Existing Wind'--------------------
if ExistingWind ==1:
    # 2 Solar
    page['E4'] = MWhS            # MWh Solar
    page['F4'] = Solar_D         # $ Solar
    page['G4'] = MWhS_Tcrt       # MWh Solar Curtailment
    page['H4'] = SolarCurt_D     # $ Solar Curtailment
    page['I4'] = Sol_DpE         # $/MWh Solar
    # 5 CC
    page['E5'] = SumMWhCC        # MWh yearly CC
    page['F5'] = CC_D            # $ yearly CC
    page['I5'] = CC_DpE          # $/MWh yearly CC
    # 6 CT
    page['E6'] = SumMWhCT        # MWh yearly CT
    page['F6'] = CT_D            # $ yearly CT
    page['I6'] = CT_DpE          # $/MWh yearly CT
    # 7 RIC
    page['E7'] = SumMWhRIC       # MWh yearly RIC
    page['F7'] = RIC_D           # $ yearly RIC
    page['I7'] = RIC_DpE         # $/MWh yearly RIC
    # 8 Aro
    page['E8'] = SumMWhA         # MWh yearly Aro
    page['F8'] = A_D             # $ yearly Aro
    page['I8'] = A_DpE           # $/MWh yearly Aro
    # 24 solar + Bat
    page['E9'] = SBE             # MWh solar + Battery
    page['F9'] = SBD             # $ Solar + Battery
    page['G9'] = SBCrE           # MWh solar Curtailment
    page['H9'] = SBCrD           # $ solar Curtailment
    page['I9'] = SBDpE           # $/MWh solar + Battery                                    
    #  25 Solar + Cc                                    
    page['E10'] = SCcE           # MWh solar + CC
    page['F10'] = SCcD           # $ Solar + CC
    page['G10'] = SCcCrE         # MWh Solar Curtailment
    page['H10'] = SCcCrD         # $ Solar Curtailment
    page['I10'] = SCcDpE         # $/MWh Solar + CC                                    
    #  26 Solar + Ct                                    
    page['E11'] = SCtE           # MWh solar + CT
    page['F11'] = SCtD           # $ Solar + CT
    page['G11'] = SCtCrE         # MWh Solar Curtailment
    page['H11'] = SCtCrD         # $ Solar Curtailment
    page['I11'] = SCtDpE         # $/MWh Solar + CT                                   
    #  27 Solar + Ric                                    
    page['E12'] = SRicE          # MWh solar + RIC
    page['F12'] = SRicD          # $ Solar + RIC
    page['G12'] = SRicCrE        # MWh Solar Curtailment
    page['H12'] = SRicCrD        # $ Solar Curtailment
    page['I12'] = SRicDpE        # $/MWh Solar + RIC                                  
    #  28 Solar + Aro (_a)                                   
    page['E13'] = S_aE           # MWh solar + Aro
    page['F13'] = S_aD           # $ Solar + Aro
    page['G13'] = S_aCrE         # MWh Solar Curtailment
    page['H13'] = S_aCrD         # $ Solar Curtailment
    page['I13'] = S_aDpE         # $/MWh Solar + Aro  
#==================Plots====================================
    #---------------matplotlib-----------------------------------
    import matplotlib.pyplot as plt
    
    # plot MWh by Resources
    plt.figure (1)
    
    names = ['Solar','CC','CT','RIC','Aro',
          'Solar\n + \nBat','Solar\n + \nCc','Solar\n + \nCt',
          'Solar\n + \nRic','Solar\n + \n_aro','S_\nCurt','s_\nCurt']
    values= [MWhS,SumMWhCC,SumMWhCT,SumMWhRIC,SumMWhA,
          SBE,SCcE,SCtE,SRicE,S_aE,MWhS_Tcrt,SBCrE]
    
    plt.title('Energy by Resources')
    plt.ylabel('Energy [MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='b')
    
    plt.show()
    #-----------------------------------------------------------
    # plot $ by Resources
    plt.figure (2)
    
    names = ['Solar','CC','CT','RIC','Aro',
          'Solar\n + \nBat','Solar\n + \nCc','Solar\n + \nCt',
          'Solar\n + \nRic','Solar\n + \n_aro','S_\nCurt','s_\nCurt']
    values= [Solar_D,CC_D,CT_D,RIC_D,A_D,
          SBD,SCcD,SCtD,SRicD,S_aD,SolarCurt_D,SBCrD]
    
    plt.title('Cost by Resources')
    plt.ylabel('dollars [$]')
    plt.xlabel('Resources')
    plt.bar(names, values, color ='r')
    
    plt.show()
    #-----------------------------------------------------------
    # plot 'Energy Price per Resource'
    plt.figure (3)
    
    names = ['Solar','CC','CT','RIC','Aro',
          'Solar\n + \nBat','Solar\n + \nCc','Solar\n + \nCt',
          'Solar\n + \nRic','Solar\n + \n_aro']
    values= [Sol_DpE,CC_DpE,CT_DpE,RIC_DpE,A_DpE,
          SBDpE,SCcDpE,SCtDpE,SRicDpE,S_aDpE]
    
    plt.title('Energy Price per Resource')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values,color ='c')
    
    plt.show()
    #--------Excel using (openpyxl)---From Existing Wind---------------------
    from openpyxl.chart import BarChart, Reference, Series
    
    # Figure 1 'Energy vs Resources'
    W_energy = Reference(page, min_row=4, max_row=13, min_col=5,max_col=5)
    W_resources = Reference (page, min_row=4, max_row=13, min_col=4,max_col=4 )
    Rplot =Series( W_energy, title= 'resources')                             
    chartWR = BarChart()
    chartWR.append(Rplot)
    chartWR.set_categories(W_resources)
    # labeling
    chartWR.title = 'Energy by Resource'
    chartWR.x_axis.title = 'Resources'
    chartWR.y_axis.title = 'Energy [MWh]'
    # chart dimentions
    chartWR.height = 10 # default is 7.5
    chartWR.width = 20  # default is 15
    # place on excel
    page.add_chart(chartWR,'K15')
      
    # Figure 2 'Dollars vs Resources'
    W_dollars = Reference(page, min_row=4, max_row=13, min_col=6,max_col=6)
    #resources = Reference (page, min_row=4, max_row=13, min_col=4,max_col=4 )
    Dplot =Series(W_dollars, title= 'resources')                             
    chartWD = BarChart()
    chartWD.append(Dplot)
    chartWD.set_categories(W_resources)
    # labeling
    chartWD.title = 'Dollars by Resource'
    chartWD.x_axis.title = 'Resources'
    chartWD.y_axis.title = 'Dollars [$]'
    # chart dimentions
    chartWD.height = 10 # default is 7.5
    chartWD.width = 20  # default is 15
    # place on excel
    page.add_chart(chartWD,'W15')
       
    # Figure 3 'Energy Price per Resource'
    W_dolpMWh = Reference(page, min_row=4, max_row=13, min_col=9,max_col=9)
    #resources = Reference (page, min_row=4, max_row=13, min_col=4,max_col=4 )
    DpEplot =Series(W_dolpMWh, title= 'resources')                             
    chartDpE = BarChart()
    chartDpE.append(DpEplot)
    chartDpE.set_categories(W_resources)
    # labeling
    chartDpE.title = 'Energy Price per Resource'
    chartDpE.x_axis.title = 'Resources'
    chartDpE.y_axis.title = 'Energy Cost [$/MWh]'
    # chart dimentions
    chartDpE.height = 10 # default is 7.5
    chartDpE.width = 20  # default is 15
    # place on excel
    page.add_chart(chartDpE,'Q4')
else:  
    # -------write on Excel Sheet 'EcoAnResulsM3.xlsx'-------------------
    #-------Resources & combinations from 'Existing Solar'--------------------
    # 1 Wind
    page['E28'] = MWhW           # MWh Wind
    page['F28'] = Wind_D         # $ Wind
    page['G28'] = MWhW_Tcrt      # MWh Wind Curtailment
    page['H28'] = WindCurt_D     # $ Wind Curtailment
    page['I28'] = W_DpE          # $/MWh Wind
    # 5 CC
    # write on Excel Sheet
    page['E29'] = SumMWhCC       # MWh yearly CC
    page['F29'] = CC_D           # $ yearly CC
    page['I29'] = CC_DpE         # $/MWh yearly CC
    # 6 CT
    page['E30'] = SumMWhCT       # MWh yearly CT
    page['F30'] = CT_D           # $ yearly CT
    page['I30'] = CT_DpE         # $/MWh yearly CT
    # 7 RIC
    # write on Excel Sheet
    page['E31'] = SumMWhRIC      # MWh yearly RIC
    page['F31'] = RIC_D          # $ yearly RIC
    page['I31'] = RIC_DpE        # $/MWh yearly RIC
    # 8 Aro
    page['E32'] = SumMWhA        # MWh yearly Aro
    page['F32'] = A_D            # $ yearly Aro
    page['I32'] = A_DpE          # $/MWh yearly Aro
    # 13 Wind + Bat
    page['E33'] = WBE            # MWh Wind + Battery
    page['F33'] = WBD            # $ Wind + Battery
    page['G33'] = WBCrE          # MWh Wind Curtailment
    page['H33'] = WBCrD          # $ Wind Curtailment
    page['I33'] = WBDpE          # $/MWh Wind + Battery                                 
    #  15 Wind + CC                                    
    page['E34'] = WCcE           # MWh Wind + CC
    page['F34'] = WCcD           # $ Wind + CC
    page['G34'] = WCcCrE         # MWh Wind Curtailment
    page['H34'] = WCcCrD         # $ Wind Curtailment
    page['I34'] = WCcDpE         # $/MWh Wind + CC                                  
    #  16 Wind + CT                                    
    page['E35'] =  WCtE          # MWh Wind + CT
    page['F35'] =  WCtD          # $ Wind + CT
    page['G35'] =  WCtCrE        # MWh Wind Curtailment
    page['H35'] =  WCtCrD        # $ Wind Curtailment
    page['I35'] =  WCtDpE        # $/MWh Wind + CT
    #  17 Wind + RIC                                    
    page['E36'] = WRicE          # MWh Wind + RIC
    page['F36'] = WRicD          # $ Wind + RIC
    page['G36'] = WRicCrE        # MWh Wind Curtailment
    page['H36'] = WRicCrD        # $ Wind Curtailment
    page['I36'] = WRicDpE        # $/MWh Wind + RIC
    #  18 Wind + Aro (_a)                                    
    page['E37'] =  W_aE          # MWh Wind + Aro
    page['F37'] =  W_aD          # $ Wind + Aro
    page['G37'] =  W_aCrE        # MWh Wind Curtailment
    page['H37'] =  W_aCrD        # $ Wind Curtailment
    page['I37'] =  W_aDpE        # $/MWh Wind + Aro  
    #==================Plots====================================
    #---------------matplotlib-----------------------------------
    import matplotlib.pyplot as plt
    
    # plot MWh by Resources
    plt.figure (1)
    
    names = ['Wind','CC','CT','RIC','Aro',
              'Wind\n + \nBat','Wind\n + \nCc','Wind\n + \nCt',
              'Wind\n + \nRic','Wind\n + \n_aro','W_\nCurt','w_\nCurt']
    values= [MWhW,SumMWhCC,SumMWhCT,SumMWhRIC,SumMWhA,
              WBE,WCcE,WCtE,WRicE,W_aE,MWhW_Tcrt,WBCrE]
    
    plt.title('Energy by Resources')
    plt.ylabel('Energy [MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='b')
    
    plt.show()
   
    #-----------------------------------------------------------
    # plot $ by Resources
    plt.figure (2)
    
    names = ['Wind','CC','CT','RIC','Aro',
              'Wind\n + \nBat','Wind\n + \nCc','Wind\n + \nCt',
              'Wind\n + \nRic','Wind\n + \n_aro','W_\nCurt','w_\nCurt']
    values= [MWhW,SumMWhCC,SumMWhCT,SumMWhRIC,SumMWhA,
              WBE,WCcE,WCtE,WRicE,W_aE,MWhW_Tcrt,WBCrE]
    
    
    plt.title('Cost by Resources')
    plt.ylabel('dollars [$]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='r')
    
    plt.show()
    #-----------------------------------------------------------
    # plot 'Energy Price per Resource'
    plt.figure (3)
    
    names = ['Wind','CC','CT','RIC','Aro',
              'Wind\n + \nBat','Wind\n + \nCc','Wind\n + \nCt',
              'Wind\n + \nRic','Wind\n + \n_aro']
    values= [W_DpE,CC_DpE,CT_DpE,RIC_DpE,A_DpE,
              WBDpE,WCcDpE,WCtDpE,WRicDpE,W_aDpE]
    
    plt.title('Energy Price per Resource')
    plt.ylabel('dollars per Energy [$/MWh]')
    plt.xlabel('Resources')
    plt.bar(names, values, color='c')
    
    plt.show()
    #--------Excel using (openpyxl)---From Existing Solar--
    from openpyxl.chart import BarChart, Reference, Series
    
    # Figure 1 'Energy vs Resources'
    S_energy = Reference(page, min_row=28, max_row=37, min_col=5,max_col=5)
    S_resources = Reference (page, min_row=28, max_row=37, min_col=4,max_col=4 )
    Rplot =Series( S_energy, title= 'resources')                             
    chartSR = BarChart()
    chartSR.append(Rplot)
    chartSR.set_categories(S_resources)
    # labeling
    chartSR.title = 'Energy by Resource'
    chartSR.x_axis.title = 'Resources'
    chartSR.y_axis.title = 'Energy [MWh]'
    # chart dimentions
    chartSR.height = 10 # default is 7.5
    chartSR.width = 20  # default is 15
    # place on excel
    page.add_chart(chartSR,'K39')
      
    # Figure 2 'Dollars vs Resources'
    S_dollars = Reference(page, min_row=28, max_row=37, min_col=6,max_col=6)
    #resources = Reference (page, min_row=4, max_row=13, min_col=4,max_col=4 )
    Dplot =Series(S_dollars, title= 'resources')                             
    chartSD = BarChart()
    chartSD.append(Dplot)
    chartSD.set_categories(S_resources)
    # labeling
    chartSD.title = 'Dollars by Resource'
    chartSD.x_axis.title = 'Resources'
    chartSD.y_axis.title = 'Dollars [$]'
    # chart dimentions
    chartSD.height = 10 # default is 7.5
    chartSD.width = 20  # default is 15
    # place on excel
    page.add_chart(chartSD,'W39')
       
    # Figure 3 'Energy Price per Resource'
    S_dolpMWh = Reference(page, min_row=28, max_row=37, min_col=9,max_col=9)
    #resources = Reference (page, min_row=4, max_row=13, min_col=4,max_col=4 )
    DpEplot =Series(S_dolpMWh, title= 'resources')                             
    chartDpE = BarChart()
    chartDpE.append(DpEplot)
    chartDpE.set_categories(S_resources)
    # labeling
    chartDpE.title = 'Energy Price per Resource'
    chartDpE.x_axis.title = 'Resources'
    chartDpE.y_axis.title = 'Energy Cost [$/MWh]'
    # chart dimentions
    chartDpE.height = 10 # default is 7.5
    chartDpE.width = 20  # default is 15
    # place on excel
    page.add_chart(chartDpE,'Q28')
    
# ===========******'Save' 'EcoAnResulsM3.xlsx'*******==========
wbd.save('EcoAnResulsM3.xlsx')




















